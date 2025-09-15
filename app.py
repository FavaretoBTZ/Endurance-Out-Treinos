import pandas as pd
import re
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import ScatterChart, Series, Reference

# Caminho para a pasta Downloads no Mac
downloads_path = os.path.expanduser("~/Downloads")

# Caminhos dos arquivos
csv_path = os.path.join(downloads_path, "TODOS - 3 TREINO LIVRE OFICIAL - GT3   GT4   P1 - laptimes.csv")
output_path = os.path.join(downloads_path, "25ET4 TL3.xlsx")

# Colunas desejadas
colunas_desejadas = ['Lap', 'LeadLap', 'Lap Tm', 'S1 Tm', 'S2 Tm', 'S3 Tm', 'SSTRAP', 'SSTRAP Tm']

# Ordem fixa das abas
ordem_fixa = [
    "MAURO-PAVIE-BOTTU - P1",
    "GORAYEB-BUZAID - P1",
    "L.FORES-MARTI-V.FORES - P1",
    "LANCAS-MORAES-ABRUNH - P1",
    "CARLESSO-MORGATTO - P1",
    "OHASHI - PADRON - P1",
    "ROCHA-ASSUNCAO - P1"
]

# ---------------- Helpers ----------------
def normalizar_nomes(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nomes comuns de colunas para evitar KeyError."""
    new_cols = []
    for c in df.columns:
        c0 = str(c).replace("\xa0", " ").strip()
        c0 = re.sub(r"\s+", " ", c0)
        if c0.lower() == "lead lap": c0 = "LeadLap"
        if c0.lower() == "lap time": c0 = "Lap Tm"
        new_cols.append(c0)
    df.columns = new_cols
    return df

def tempo_para_segundos(t):
    try:
        ts = str(t).replace(',', '.').strip()
        if ':' in ts:
            m, s = ts.split(':', 1)
            return int(m) * 60 + float(s)
        return float(ts)
    except:
        return None

def extrair_nome_treino_de_arquivo(path: str) -> str:
    """Extrai o nome do treino a partir do nome do arquivo."""
    base = os.path.basename(path)
    m = re.search(r'-\s*([^-]+?)\s*-\s*', base)
    if m:
        return m.group(1).strip()
    partes = base.split(" - ")
    if len(partes) > 1:
        return partes[1].strip()
    return "TREINO"

def nome_aba_excel(nome: str) -> str:
    """Sanitiza o nome da aba para ser válido no Excel."""
    return re.sub(r'[\\/*?:\[\]]', '-', nome)[:31]

def mapear_nome(nome_original: str) -> str:
    """Mapeia nomes originais do CSV para os nomes padronizados usando palavras-chave."""
    nome_original = nome_original.upper()

    if "MAURO" in nome_original and "PAVIE" in nome_original:
        return "MAURO-PAVIE-BOTTU - P1"
    if "GORAYEB" in nome_original or "BUZAID" in nome_original:
        return "GORAYEB-BUZAID - P1"
    if "FORES" in nome_original:
        return "L.FORES-MARTI-V.FORES - P1"
    if "LANCAS" in nome_original or "MORAES" in nome_original or "ABRUNH" in nome_original:
        return "LANCAS-MORAES-ABRUNH - P1"
    if "CARLESSO" in nome_original or "MORGATTO" in nome_original:
        return "CARLESSO-MORGATTO - P1"
    if "OHASHI" in nome_original or "PADRON" in nome_original:
        return "OHASHI - PADRON - P1"
    if "ROCHA" in nome_original or "ASSUNCAO" in nome_original:
        return "ROCHA-ASSUNCAO - P1"

    return nome_original  # fallback, caso não bata em nenhum

# ---------------- Leitura do CSV ----------------
csv_df = pd.read_csv(csv_path)
csv_df = normalizar_nomes(csv_df)

# Nome do treino
treino_nome = extrair_nome_treino_de_arquivo(csv_path)
print(f"Treino identificado: {treino_nome}")

# Separar dados por carro
dados_por_carro = {}
current_car = None
for _, row in csv_df.iterrows():
    if pd.isna(row.get('Lap')) and isinstance(row.get('Time of Day'), str):
        current_car = row['Time of Day'].strip()
        current_car = mapear_nome(current_car)  # 🔹 aplica mapeamento flexível
        dados_por_carro[current_car] = []
    elif current_car:
        dados_por_carro[current_car].append(row)

# Criar DataFrames por carro
abas_temp = {nome: pd.DataFrame(linhas) for nome, linhas in dados_por_carro.items()}

# Reorganizar de acordo com a ordem fixa
abas = {}
for nome in ordem_fixa:
    if nome in abas_temp:
        abas[nome] = abas_temp[nome]
    else:
        # cria aba vazia só com os títulos
        abas[nome] = pd.DataFrame(columns=colunas_desejadas)

# ---------------- Gravar Excel ----------------
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for nome_aba, df in abas.items():
        df = normalizar_nomes(df)
        cols_presentes = [c for c in colunas_desejadas if c in df.columns]
        df_filtrado = df[cols_presentes].copy()

        # Adiciona coluna "Treino"
        df_filtrado["Treino"] = treino_nome if not df_filtrado.empty else None

        # 🔹 Nome seguro para Excel
        nome_excel = nome_aba_excel(nome_aba)

        df_filtrado.to_excel(writer, sheet_name=nome_excel, index=False)

# ---------------- Pós-processamento no Excel ----------------
wb = load_workbook(output_path)

# Criar/Recriar aba "Base Gráfico"
if "Base Gráfico" in wb.sheetnames:
    del wb["Base Gráfico"]
ws_base = wb.create_sheet(title="Base Gráfico")
ws_base.append(["Carro", "LeadLap", "Lap Tm (s)"])

for nome_aba, df in abas.items():
    df = normalizar_nomes(df)
    if not {'LeadLap', 'Lap Tm'}.issubset(df.columns):
        continue
    df_filtrado = df[['LeadLap', 'Lap Tm']].dropna(subset=['LeadLap', 'Lap Tm']).copy()
    df_filtrado["Lap Tm (s)"] = df_filtrado["Lap Tm"].apply(tempo_para_segundos)
    df_filtrado = df_filtrado.dropna(subset=["Lap Tm (s)"])
    for _, row in df_filtrado.iterrows():
        ws_base.append([nome_aba_excel(nome_aba), row["LeadLap"], row["Lap Tm (s)"]])

# Criar gráfico
chart = ScatterChart()
chart.title = "Lap Tm vs LeadLap por Carro"
chart.x_axis.title = "LeadLap"
chart.y_axis.title = "Lap Tm (s)"
chart.legend.position = "r"
chart.style = 2

linha_total = ws_base.max_row
carros = sorted(list(set(ws_base.cell(row=r, column=1).value for r in range(2, linha_total + 1))))
for carro in carros:
    idx = [r for r in range(2, linha_total + 1) if ws_base.cell(row=r, column=1).value == carro]
    if not idx:
        continue
    x_ref = Reference(ws_base, min_col=2, min_row=idx[0], max_row=idx[-1])
    y_ref = Reference(ws_base, min_col=3, min_row=idx[0], max_row=idx[-1])
    chart.series.append(Series(y_ref, x_ref, title=carro[:31]))
ws_base.add_chart(chart, "E2")

# Salvar final
wb.save(output_path)
print(f"Arquivo Excel salvo em: {output_path}")
